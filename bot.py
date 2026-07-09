# -*- coding: utf-8 -*-
"""
Telegram-бот для сбора ежедневных отчётов менеджеров GET IT.
Роли, поля и расписание — в config.py.

Запуск: python bot.py
"""
import asyncio
import io
import json
import logging
import os
from datetime import datetime, date as date_cls, timedelta

from aiogram import Bot, Dispatcher, Router, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton,
    BufferedInputFile,
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config
import db

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("partner_bot")

bot = Bot(token=config.BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.MARKDOWN))
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)


# ============================================================================
# FSM состояния
# ============================================================================
class Registration(StatesGroup):
    waiting_name = State()
    waiting_role = State()


class Survey(StatesGroup):
    filling = State()
    waiting_manual = State()


class SetPlan(StatesGroup):
    waiting_manager = State()
    waiting_field = State()
    waiting_value = State()


class SetPlanAll(StatesGroup):
    waiting_field = State()
    waiting_value = State()


# ============================================================================
# Вспомогательные функции
# ============================================================================
def is_admin(telegram_id: int) -> bool:
    return telegram_id == config.ADMIN_ID


def today_str() -> str:
    return date_cls.today().isoformat()


def indicator(fact: float, plan: float) -> str:
    """Цветовая индикация выполнения плана."""
    if not plan or plan <= 0:
        return "⬜"
    pct = fact / plan * 100
    if pct >= 100:
        return "🟢"
    if pct >= 70:
        return "🟡"
    return "🔴"


def fmt_value(value: float, field_type: str) -> str:
    if field_type == "money":
        return f"{int(value):,}".replace(",", " ")
    if value == int(value):
        return str(int(value))
    return str(value)


def roles_keyboard() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    for role in config.ROLES:
        kb.button(text=role, callback_data=f"role:{role}")
    kb.adjust(1)
    return kb.as_markup()


def quick_values_keyboard(field: dict) -> InlineKeyboardMarkup:
    values = config.QUICK_MONEY_VALUES if field["type"] == "money" else config.QUICK_COUNT_VALUES
    kb = InlineKeyboardBuilder()
    for v in values:
        kb.button(text=fmt_value(v, field["type"]), callback_data=f"qv:{v}")
    kb.adjust(5 if field["type"] == "count" else 3)
    kb.row(InlineKeyboardButton(text="✏️ Ввести вручную", callback_data="qv:manual"))
    kb.row(InlineKeyboardButton(text="⏭ Пропустить (0)", callback_data="qv:0"))
    return kb.as_markup()


def field_prompt_text(field: dict, step: int, total: int) -> str:
    return (
        f"Шаг {step}/{total}\n\n"
        f"{field['emoji']} *{field['label']}*\n"
        f"Введите значение ({field['unit']}) или выберите кнопкой ниже:"
    )


async def send_to_groups(text: str) -> None:
    """Отправить текст во все настроенные группы (id == 0 пропускается)."""
    for gid in config.GROUP_IDS:
        if not gid:
            continue
        try:
            await bot.send_message(gid, text)
        except Exception as e:
            log.warning("Не удалось отправить сообщение в группу %s: %s", gid, e)


# ============================================================================
# /start — регистрация менеджера
# ============================================================================
@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    manager = db.get_manager_by_telegram_id(message.from_user.id)
    if manager and not manager["active"]:
        await message.answer(
            "Ваш профиль сейчас отключён администратором.\n"
            "Если это ошибка — напишите ему напрямую."
        )
        return
    if manager:
        await message.answer(
            f"С возвращением, *{manager['name']}*! 👋\n"
            f"Роль: {manager['role']}\n\n"
            f"Команды:\n"
            f"/report — заполнить отчёт за сегодня\n"
            f"/my — мои данные\n"
            f"/help — справка"
        )
        return

    await state.set_state(Registration.waiting_name)
    await message.answer(
        "👋 Привет! Рад видеть в строю. Я собираю ежедневные отчёты.\n"
        "Напиши свои имя и фамилию, чтобы я закрепил за тобой профиль 👇"
    )


@router.message(Registration.waiting_name)
async def reg_name(message: Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("Имя слишком короткое, попробуйте ещё раз 🙂")
        return
    await state.update_data(name=name)
    await state.set_state(Registration.waiting_role)
    await message.answer("Отлично! Теперь выберите вашу роль:", reply_markup=roles_keyboard())


@router.callback_query(Registration.waiting_role, F.data.startswith("role:"))
async def reg_role(callback: CallbackQuery, state: FSMContext):
    role = callback.data.split(":", 1)[1]
    data = await state.get_data()
    name = data["name"]
    manager_id = db.generate_manager_id(name)

    db.add_manager(
        telegram_id=callback.from_user.id,
        manager_id=manager_id,
        name=name,
        role=role,
        username=callback.from_user.username,
    )
    await state.clear()

    await callback.message.edit_text(
        f"✅ Регистрация завершена!\n\n"
        f"Имя: *{name}*\nРоль: *{role}*\n\n"
        f"Каждый рабочий день в {config.SURVEY_HOUR}:{config.SURVEY_MINUTE:02d} я буду напоминать "
        f"заполнить отчёт. Заполнить можно и вручную командой /report."
    )
    await callback.answer()

    if config.ADMIN_ID:
        try:
            await bot.send_message(
                config.ADMIN_ID,
                f"🆕 Новая регистрация в боте {config.COMPANY_NAME}:\n"
                f"*{name}* — {role}\n"
                f"(@{callback.from_user.username or 'без username'})"
            )
        except Exception as e:
            log.warning("Не удалось уведомить админа о регистрации: %s", e)


# ============================================================================
# /report — пошаговый опросник
# ============================================================================
async def ask_step(target: Message, state: FSMContext, edit: bool = False):
    data = await state.get_data()
    step = data["step"]
    total = len(config.FIELDS)
    field = config.FIELDS[step]
    text = field_prompt_text(field, step + 1, total)
    kb = quick_values_keyboard(field)
    if edit:
        await target.edit_text(text, reply_markup=kb)
    else:
        await target.answer(text, reply_markup=kb)


@router.message(Command("report"))
async def cmd_report(message: Message, state: FSMContext):
    manager = db.get_manager_by_telegram_id(message.from_user.id)
    if not manager:
        await message.answer("Вы ещё не зарегистрированы. Нажмите /start, чтобы начать.")
        return
    if not manager["active"]:
        await message.answer("Ваш профиль отключён администратором — отчёт временно недоступен.")
        return

    date = today_str()
    db.start_survey(message.from_user.id, manager["manager_id"], date)
    await state.set_state(Survey.filling)
    await state.update_data(manager_id=manager["manager_id"], date=date, step=0, answers={})

    await message.answer(f"📋 Заполняем отчёт за {date}. Отвечайте по одному пункту.")
    await ask_step(message, state)


async def _save_answer_and_advance(event, state: FSMContext, value: float):
    data = await state.get_data()
    step = data["step"]
    manager_id = data["manager_id"]
    date = data["date"]
    field = config.FIELDS[step]
    answers = data["answers"]
    answers[field["key"]] = value

    db.upsert_activity(manager_id, date, field["key"], fact=value)

    step += 1
    db.update_survey_state(event.from_user.id, step, answers)

    if step >= len(config.FIELDS):
        await state.clear()
        db.clear_survey_state(event.from_user.id)
        summary_lines = []
        for f in config.FIELDS:
            v = answers.get(f["key"], 0)
            summary_lines.append(f"{f['emoji']} {f['label']}: *{fmt_value(v, f['type'])}* {f['unit']}")
        text = "✅ Отчёт за сегодня сохранён!\n\n" + "\n".join(summary_lines)
        if isinstance(event, CallbackQuery):
            await event.message.edit_text(text)
        else:
            await event.answer(text)
        return

    await state.update_data(step=step, answers=answers)
    if isinstance(event, CallbackQuery):
        await ask_step(event.message, state, edit=True)
    else:
        await ask_step(event, state)


@router.callback_query(Survey.filling, F.data.startswith("qv:"))
async def survey_quick_value(callback: CallbackQuery, state: FSMContext):
    raw = callback.data.split(":", 1)[1]
    await callback.answer()
    if raw == "manual":
        await state.set_state(Survey.waiting_manual)
        data = await state.get_data()
        field = config.FIELDS[data["step"]]
        await callback.message.edit_text(
            f"{field['emoji']} *{field['label']}*\nВведите число ({field['unit']}):"
        )
        return
    value = float(raw)
    await _save_answer_and_advance(callback, state, value)


@router.message(Survey.waiting_manual)
async def survey_manual_value(message: Message, state: FSMContext):
    text = message.text.strip().replace(" ", "").replace(",", ".")
    try:
        value = float(text)
        if value < 0:
            raise ValueError
    except ValueError:
        await message.answer("Нужно положительное число. Попробуйте ещё раз:")
        return
    await state.set_state(Survey.filling)
    await _save_answer_and_advance(message, state, value)


# ============================================================================
# /my — данные менеджера за день и месяц
# ============================================================================
@router.message(Command("my"))
async def cmd_my(message: Message):
    manager = db.get_manager_by_telegram_id(message.from_user.id)
    if not manager:
        await message.answer("Вы ещё не зарегистрированы. Нажмите /start.")
        return

    date = today_str()
    today_data = db.get_activities_for_date(date).get(manager["manager_id"], {})

    first_day = date_cls.today().replace(day=1).isoformat()
    month_data = db.get_activities_for_period(first_day, date).get(manager["manager_id"], {})

    lines = [f"📊 *{manager['name']}* ({manager['role']})\n", f"*Сегодня, {date}:*"]
    for f in config.FIELDS:
        d = today_data.get(f["key"], {"plan": 0, "fact": 0})
        ind = indicator(d["fact"], d["plan"])
        lines.append(f"{ind} {f['emoji']} {f['label']}: {fmt_value(d['fact'], f['type'])} / план {fmt_value(d['plan'], f['type'])} {f['unit']}")

    lines.append(f"\n*С начала месяца:*")
    for f in config.FIELDS:
        d = month_data.get(f["key"], {"plan": 0, "fact": 0})
        ind = indicator(d["fact"], d["plan"])
        lines.append(f"{ind} {f['emoji']} {f['label']}: {fmt_value(d['fact'], f['type'])} / план {fmt_value(d['plan'], f['type'])} {f['unit']}")

    await message.answer("\n".join(lines))


# ============================================================================
# /help
# ============================================================================
@router.message(Command("help"))
async def cmd_help(message: Message):
    manager = db.get_manager_by_telegram_id(message.from_user.id)
    text = (
        f"🤖 *Бот ежедневных отчётов — {config.COMPANY_NAME}*\n\n"
        "*Команды менеджера:*\n"
        "/start — регистрация\n"
        "/report — заполнить отчёт за сегодня\n"
        "/my — мои данные за день и месяц\n"
        "/help — эта справка\n"
    )
    if is_admin(message.from_user.id):
        text += (
            "\n*Команды администратора:*\n"
            "/summary — сводка за сегодня\n"
            "/month — итоги за месяц\n"
            "/send_summary — отправить сводку в группы сейчас\n"
            "/send_month — отправить месячную сводку в группы\n"
            "/team — список менеджеров\n"
            "/setplan — установить план одному менеджеру\n"
            "/setplan_all — установить одинаковый план всем менеджерам\n"
            "/remove — убрать менеджера из отчётов и сводок\n"
            "/activate — вернуть ранее убранного менеджера\n"
            "/export — выгрузить JSON за месяц\n"
            "/export_excel — выгрузить Excel за месяц (или `/export_excel 2026-07`)\n"
            "/backup — скачать полную копию базы данных\n"
            "/chatid — узнать ID текущего чата\n"
        )
    if not manager and not is_admin(message.from_user.id):
        text += "\nЧтобы начать — нажмите /start."
    await message.answer(text)


# ============================================================================
# Формирование сводок
# ============================================================================
def build_daily_summary(date: str) -> str:
    managers = db.get_all_managers()
    activities = db.get_activities_for_date(date)

    reported = sum(1 for m in managers if db.has_reported(m["manager_id"], date))
    lines = [f"📊 *Сводка {config.COMPANY_NAME} за {date}*", f"Заполнили: {reported}/{len(managers)}\n"]

    if not managers:
        lines.append("Пока нет зарегистрированных менеджеров.")
        return "\n".join(lines)

    for m in managers:
        data = activities.get(m["manager_id"], {})
        lines.append(f"👤 *{m['name']}* ({m['role']})")
        for f in config.FIELDS:
            d = data.get(f["key"], {"plan": 0, "fact": 0})
            ind = indicator(d["fact"], d["plan"])
            lines.append(f"  {ind} {f['emoji']} {f['label']}: {fmt_value(d['fact'], f['type'])}/{fmt_value(d['plan'], f['type'])} {f['unit']}")
        lines.append("")

    return "\n".join(lines).strip()


def build_monthly_summary(year: int, month: int) -> str:
    start = date_cls(year, month, 1).isoformat()
    if month == 12:
        end = date_cls(year, 12, 31).isoformat()
    else:
        end = (date_cls(year, month + 1, 1) - timedelta(days=1)).isoformat()

    managers = db.get_all_managers()
    activities = db.get_activities_for_period(start, end)

    lines = [f"📈 *Итоги {config.COMPANY_NAME} за {start[:7]}*\n"]
    if not managers:
        lines.append("Пока нет зарегистрированных менеджеров.")
        return "\n".join(lines)

    for m in managers:
        data = activities.get(m["manager_id"], {})
        lines.append(f"👤 *{m['name']}* ({m['role']})")
        for f in config.FIELDS:
            d = data.get(f["key"], {"plan": 0, "fact": 0})
            ind = indicator(d["fact"], d["plan"])
            lines.append(f"  {ind} {f['emoji']} {f['label']}: {fmt_value(d['fact'], f['type'])}/{fmt_value(d['plan'], f['type'])} {f['unit']}")
        lines.append("")

    return "\n".join(lines).strip()


# ============================================================================
# Админские команды
# ============================================================================
@router.message(Command("summary"))
async def cmd_summary(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer(build_daily_summary(today_str()))


@router.message(Command("month"))
async def cmd_month(message: Message):
    if not is_admin(message.from_user.id):
        return
    today = date_cls.today()
    await message.answer(build_monthly_summary(today.year, today.month))


@router.message(Command("send_summary"))
async def cmd_send_summary(message: Message):
    if not is_admin(message.from_user.id):
        return
    text = build_daily_summary(today_str())
    await bot.send_message(config.ADMIN_ID, text)
    await send_to_groups(text)
    await message.answer("✅ Сводка отправлена в группы.")


@router.message(Command("send_month"))
async def cmd_send_month(message: Message):
    if not is_admin(message.from_user.id):
        return
    today = date_cls.today()
    text = build_monthly_summary(today.year, today.month)
    await bot.send_message(config.ADMIN_ID, text)
    await send_to_groups(text)
    await message.answer("✅ Месячная сводка отправлена в группы.")


@router.message(Command("team"))
async def cmd_team(message: Message):
    if not is_admin(message.from_user.id):
        return
    managers = db.get_all_managers(active_only=False)
    if not managers:
        await message.answer("Пока никто не зарегистрирован.")
        return
    lines = ["👥 *Зарегистрированные менеджеры:*\n"]
    for m in managers:
        status = "✅" if m["active"] else "🚫"
        lines.append(f"{status} *{m['name']}* — {m['role']} (id: `{m['manager_id']}`)")
    await message.answer("\n".join(lines))


# --- /remove: деактивировать менеджера (мягкое удаление) -------------------
@router.message(Command("remove"))
async def cmd_remove(message: Message):
    if not is_admin(message.from_user.id):
        return
    managers = db.get_all_managers(active_only=True)
    if not managers:
        await message.answer("Нет активных менеджеров для удаления.")
        return
    kb = InlineKeyboardBuilder()
    for m in managers:
        kb.button(text=f"{m['name']} ({m['role']})", callback_data=f"rm_mgr:{m['manager_id']}")
    kb.adjust(1)
    await message.answer("Кого убрать из отчётов и сводок?", reply_markup=kb.as_markup())


@router.callback_query(F.data.startswith("rm_mgr:"))
async def remove_select(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    manager_id = callback.data.split(":", 1)[1]
    manager = db.get_manager_by_manager_id(manager_id)
    if not manager:
        await callback.message.edit_text("Менеджер не найден — возможно, уже удалён.")
        await callback.answer()
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Да, убрать", callback_data=f"rm_confirm:{manager_id}")
    kb.button(text="✖️ Отмена", callback_data="rm_cancel")
    kb.adjust(2)
    await callback.message.edit_text(
        f"Убрать *{manager['name']}* ({manager['role']}) из опросника и сводок?\n"
        f"Его прошлые отчёты сохранятся — их можно будет вернуть командой /activate.",
        reply_markup=kb.as_markup(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("rm_confirm:"))
async def remove_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    manager_id = callback.data.split(":", 1)[1]
    manager = db.get_manager_by_manager_id(manager_id)
    db.set_manager_active(manager_id, False)
    name = manager["name"] if manager else manager_id
    await callback.message.edit_text(f"🚫 *{name}* убран(а) из опросника и сводок.")
    await callback.answer()


@router.callback_query(F.data == "rm_cancel")
async def remove_cancel(callback: CallbackQuery):
    await callback.message.edit_text("Отменено, никого не трогали.")
    await callback.answer()


# --- /activate: вернуть ранее удалённого менеджера --------------------------
@router.message(Command("activate"))
async def cmd_activate(message: Message):
    if not is_admin(message.from_user.id):
        return
    all_managers = db.get_all_managers(active_only=False)
    inactive = [m for m in all_managers if not m["active"]]
    if not inactive:
        await message.answer("Нет убранных менеджеров — все активны.")
        return
    kb = InlineKeyboardBuilder()
    for m in inactive:
        kb.button(text=f"{m['name']} ({m['role']})", callback_data=f"act_mgr:{m['manager_id']}")
    kb.adjust(1)
    await message.answer("Кого вернуть в опросник и сводки?", reply_markup=kb.as_markup())


@router.callback_query(F.data.startswith("act_mgr:"))
async def activate_manager(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    manager_id = callback.data.split(":", 1)[1]
    manager = db.get_manager_by_manager_id(manager_id)
    db.set_manager_active(manager_id, True)
    name = manager["name"] if manager else manager_id
    await callback.message.edit_text(f"✅ *{name}* снова в опроснике и сводках.")
    await callback.answer()


@router.message(Command("chatid"))
async def cmd_chatid(message: Message):
    await message.answer(f"ID этого чата: `{message.chat.id}`")


@router.message(Command("export"))
async def cmd_export(message: Message):
    if not is_admin(message.from_user.id):
        return
    today = date_cls.today()
    start = today.replace(day=1).isoformat()
    end = today.isoformat()
    managers = db.get_all_managers(active_only=False)
    activities = db.get_activities_for_period(start, end)

    payload = {
        "period": {"start": start, "end": end},
        "managers": [
            {
                "manager_id": m["manager_id"],
                "name": m["name"],
                "role": m["role"],
                "active": bool(m["active"]),
                "fields": activities.get(m["manager_id"], {}),
            }
            for m in managers
        ],
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)

    if len(text) > 4000:
        file = BufferedInputFile(text.encode("utf-8"), filename=f"export_{end}.json")
        await message.answer_document(file, caption="📦 Экспорт данных (JSON)")
    else:
        await message.answer(f"```json\n{text}\n```")


@router.message(Command("export_excel"))
async def cmd_export_excel(message: Message, command: CommandObject):
    if not is_admin(message.from_user.id):
        return

    today = date_cls.today()
    if command.args:
        try:
            year_str, month_str = command.args.strip().split("-")
            year, month = int(year_str), int(month_str)
        except (ValueError, AttributeError):
            await message.answer("Формат: `/export_excel 2026-07` (год-месяц). Без аргумента — текущий месяц.")
            return
    else:
        year, month = today.year, today.month

    start = date_cls(year, month, 1).isoformat()
    if month == 12:
        end = date_cls(year, 12, 31).isoformat()
    else:
        end = (date_cls(year, month + 1, 1) - timedelta(days=1)).isoformat()

    rows = db.get_raw_activities(start, end)
    if not rows:
        await message.answer(f"Нет данных за {start[:7]}.")
        return

    managers = {m["manager_id"]: m for m in db.get_all_managers(active_only=False)}

    # Сводим плоские записи в таблицу (дата, менеджер) -> {поле: (план, факт)}
    pivot: dict = {}
    for r in rows:
        key = (r["date"], r["manager_id"])
        pivot.setdefault(key, {})[r["field_key"]] = (r["plan"], r["fact"])

    wb = Workbook()
    ws = wb.active
    ws.title = start[:7]

    header = ["Дата", "Менеджер", "Роль"]
    for f in config.FIELDS:
        header.append(f"{f['label']} (факт)")
        header.append(f"{f['label']} (план)")
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for date, manager_id in sorted(pivot.keys()):
        m = managers.get(manager_id)
        name = m["name"] if m else manager_id
        role = m["role"] if m else ""
        row = [date, name, role]
        field_data = pivot[(date, manager_id)]
        for f in config.FIELDS:
            plan, fact = field_data.get(f["key"], (0, 0))
            row.append(fact)
            row.append(plan)
        ws.append(row)

    widths = [12, 24, 22] + [18] * (2 * len(config.FIELDS))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"activities_{start[:7]}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"📊 Активности за {start[:7]} — одна строка на менеджера в день, готово для сводной таблицы в Excel.",
    )


@router.message(Command("backup"))
async def cmd_backup(message: Message):
    if not is_admin(message.from_user.id):
        return
    if not os.path.exists(config.DB_PATH):
        await message.answer("Файл базы данных не найден.")
        return
    with open(config.DB_PATH, "rb") as fh:
        data = fh.read()
    filename = f"backup_{date_cls.today().isoformat()}.db"
    await message.answer_document(
        BufferedInputFile(data, filename=filename),
        caption=(
            "📦 Полный бэкап базы данных (файл SQLite).\n"
            "Храните в надёжном месте — из него можно восстановить всех менеджеров и все отчёты."
        ),
    )


# --- /setplan: FSM админа --------------------------------------------------
@router.message(Command("setplan"))
async def cmd_setplan(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    managers = db.get_all_managers()
    if not managers:
        await message.answer("Пока нет зарегистрированных менеджеров.")
        return
    kb = InlineKeyboardBuilder()
    for m in managers:
        kb.button(text=m["name"], callback_data=f"sp_mgr:{m['manager_id']}")
    kb.adjust(1)
    await state.set_state(SetPlan.waiting_manager)
    await message.answer("Кому ставим план?", reply_markup=kb.as_markup())


@router.callback_query(SetPlan.waiting_manager, F.data.startswith("sp_mgr:"))
async def setplan_manager(callback: CallbackQuery, state: FSMContext):
    manager_id = callback.data.split(":", 1)[1]
    await state.update_data(manager_id=manager_id)
    kb = InlineKeyboardBuilder()
    for f in config.FIELDS:
        kb.button(text=f"{f['emoji']} {f['label']}", callback_data=f"sp_field:{f['key']}")
    kb.adjust(1)
    await state.set_state(SetPlan.waiting_field)
    await callback.message.edit_text("По какому полю ставим план?", reply_markup=kb.as_markup())
    await callback.answer()


@router.callback_query(SetPlan.waiting_field, F.data.startswith("sp_field:"))
async def setplan_field(callback: CallbackQuery, state: FSMContext):
    field_key = callback.data.split(":", 1)[1]
    field = config.field_by_key(field_key)
    await state.update_data(field_key=field_key)
    await state.set_state(SetPlan.waiting_value)
    await callback.message.edit_text(f"Введите значение плана для «{field['label']}» ({field['unit']}):")
    await callback.answer()


@router.message(SetPlan.waiting_value)
async def setplan_value(message: Message, state: FSMContext):
    text = message.text.strip().replace(" ", "").replace(",", ".")
    try:
        value = float(text)
        if value < 0:
            raise ValueError
    except ValueError:
        await message.answer("Нужно положительное число. Попробуйте ещё раз:")
        return

    data = await state.get_data()
    manager_id = data["manager_id"]
    field_key = data["field_key"]
    field = config.field_by_key(field_key)
    date = today_str()

    days = db.set_plan(manager_id, field_key, value, date=date)
    manager = db.get_manager_by_manager_id(manager_id)
    await state.clear()
    await message.answer(
        f"✅ План установлен: *{manager['name']}* — {field['label']}: {fmt_value(value, field['type'])} {field['unit']}\n"
        f"Проставлен на {days} дн. — до конца текущего месяца. В начале следующего месяца план нужно поставить снова."
    )


# --- /setplan_all: тот же план сразу всем менеджерам ------------------------
@router.message(Command("setplan_all"))
async def cmd_setplan_all(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    managers = db.get_all_managers()
    if not managers:
        await message.answer("Пока нет зарегистрированных менеджеров.")
        return
    kb = InlineKeyboardBuilder()
    for f in config.FIELDS:
        kb.button(text=f"{f['emoji']} {f['label']}", callback_data=f"spa_field:{f['key']}")
    kb.adjust(1)
    await state.set_state(SetPlanAll.waiting_field)
    await message.answer(
        f"Ставим одинаковый план на *всех* активных менеджеров ({len(managers)} чел.).\nПо какому полю?",
        reply_markup=kb.as_markup(),
    )


@router.callback_query(SetPlanAll.waiting_field, F.data.startswith("spa_field:"))
async def setplan_all_field(callback: CallbackQuery, state: FSMContext):
    field_key = callback.data.split(":", 1)[1]
    field = config.field_by_key(field_key)
    await state.update_data(field_key=field_key)
    await state.set_state(SetPlanAll.waiting_value)
    await callback.message.edit_text(f"Введите значение плана для «{field['label']}» ({field['unit']}) на каждого:")
    await callback.answer()


@router.message(SetPlanAll.waiting_value)
async def setplan_all_value(message: Message, state: FSMContext):
    text = message.text.strip().replace(" ", "").replace(",", ".")
    try:
        value = float(text)
        if value < 0:
            raise ValueError
    except ValueError:
        await message.answer("Нужно положительное число. Попробуйте ещё раз:")
        return

    data = await state.get_data()
    field_key = data["field_key"]
    field = config.field_by_key(field_key)
    date = today_str()

    managers = db.get_all_managers()
    days = 0
    for m in managers:
        days = db.set_plan(m["manager_id"], field_key, value, date=date)
    await state.clear()
    await message.answer(
        f"✅ План «{field['label']}»: {fmt_value(value, field['type'])} {field['unit']} "
        f"установлен для всех {len(managers)} менеджеров, на {days} дн. до конца месяца."
    )


# ============================================================================
# Планировщик: напоминание об опроснике и рассылка сводки
# ============================================================================
async def scheduled_survey_reminder():
    managers = db.get_all_managers()
    date = today_str()
    for m in managers:
        if db.has_reported(m["manager_id"], date):
            continue
        try:
            await bot.send_message(
                m["telegram_id"],
                f"⏰ Время заполнить дневной отчёт!\nНажмите /report, чтобы начать."
            )
        except Exception as e:
            log.warning("Не удалось отправить напоминание менеджеру %s: %s", m["manager_id"], e)


async def scheduled_daily_summary():
    text = build_daily_summary(today_str())
    if config.ADMIN_ID:
        try:
            await bot.send_message(config.ADMIN_ID, text)
        except Exception as e:
            log.warning("Не удалось отправить сводку админу: %s", e)
    await send_to_groups(text)


def setup_scheduler() -> AsyncIOScheduler:
    scheduler = AsyncIOScheduler(timezone=config.TIMEZONE)
    scheduler.add_job(
        scheduled_survey_reminder,
        CronTrigger(day_of_week=config.WORKDAYS_CRON, hour=config.SURVEY_HOUR, minute=config.SURVEY_MINUTE),
        id="survey_reminder",
    )
    scheduler.add_job(
        scheduled_daily_summary,
        CronTrigger(day_of_week=config.WORKDAYS_CRON, hour=config.SUMMARY_HOUR, minute=config.SUMMARY_MINUTE),
        id="daily_summary",
    )
    return scheduler


# ============================================================================
# Запуск
# ============================================================================
async def main():
    if not config.BOT_TOKEN:
        raise RuntimeError("Не задан BOT_TOKEN в переменных окружения!")

    db.init_db()
    scheduler = setup_scheduler()
    scheduler.start()

    log.info("Бот запущен. Напоминания в %02d:%02d, сводки в %02d:%02d (МСК).",
              config.SURVEY_HOUR, config.SURVEY_MINUTE, config.SUMMARY_HOUR, config.SUMMARY_MINUTE)

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
